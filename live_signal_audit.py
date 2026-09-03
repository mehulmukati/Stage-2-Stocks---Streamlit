"""Shared Live Signal audit tables and Excel workbook generation.

The audit consumes the same ``_run_signal`` result rendered by the Streamlit
tab.  It never recalculates portfolio selection, so the UI, CLI and workbook
cannot silently use different rebalance implementations.
"""

from __future__ import annotations

import io
import json
import math
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from data import load_nse_holidays

_NAVY = "172554"
_BLUE = "2563EB"
_LIGHT_BLUE = "DBEAFE"
_LIGHT_GRAY = "F3F4F6"
_GREEN = "DCFCE7"
_GREEN_TEXT = "166534"
_AMBER = "FEF3C7"
_AMBER_TEXT = "92400E"
_RED = "FEE2E2"
_RED_TEXT = "991B1B"
_WHITE = "FFFFFF"
_THIN_GRAY = Side(style="thin", color="D1D5DB")

_CONFIGURATION_DATE_PARAMETERS = {
    "Signal date",
    "Portfolio start date",
}
_CONFIGURATION_INTEGER_PARAMETERS = {
    "M (entry)",
    "N (exit)",
    "Stage 2 entry threshold",
    "Stage 2 drop threshold",
    "Minimum history (sessions)",
    "Maximum circuits (1 year)",
}
_CONFIGURATION_PERCENT_PARAMETERS = {
    "Max position (%)",
    "Minimum annual return (%)",
    "Within 52-week high (%)",
    "Positive days 3M (%)",
    "Positive days 6M (%)",
    "Positive days 12M (%)",
    "Transaction cost (%)",
    "STCG rate (%)",
    "LTCG rate (%)",
}


def _variant_keys(variant: str) -> tuple[str, str, str]:
    if "Prop" in variant:
        return "prop_weights", "pre_rebalance_prop_weights", "prop_turnover_pct"
    if "Marginal" in variant:
        return "marg_weights", "pre_rebalance_marg_weights", "marg_turnover_pct"
    return "full_weights", "pre_rebalance_full_weights", "full_turnover_pct"


def _next_nse_session(value: object, holidays: set[str]) -> date:
    current = pd.Timestamp(value).date() + timedelta(days=1)
    while current.weekday() >= 5 or current.isoformat() in holidays:
        current += timedelta(days=1)
    return current


def _reason_map(values: list[str] | None) -> dict[str, str]:
    parsed: dict[str, str] = {}
    for value in values or []:
        text = str(value).strip()
        if "(" in text and text.endswith(")"):
            split_at = text.rfind("(")
            parsed[text[:split_at].strip()] = text[split_at + 1 : -1]
        else:
            parsed[text] = ""
    return parsed


def _serialise(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (pd.Timestamp, datetime, date)):
        timestamp = pd.Timestamp(value)
        if timestamp.tzinfo is not None:
            timestamp = timestamp.tz_localize(None)
        return timestamp.to_pydatetime()
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True, default=str)
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def build_live_signal_audit_tables(
    params: dict,
    result: dict,
    reconciliation: dict | None = None,
    replay_checks: list[dict] | None = None,
) -> dict[str, pd.DataFrame]:
    """Build the canonical tables used by all Live Signal audit outputs."""
    weight_key, pre_weight_key, turnover_key = _variant_keys(str(params.get("variant", "")))
    holidays = set(load_nse_holidays())
    reset_date = result.get("portfolio_reset_date")
    all_events = result.get("holdings_log", [])
    events = [
        event for event in all_events if reset_date is None or pd.Timestamp(event["date"]) >= pd.Timestamp(reset_date)
    ]

    weekly_rows: list[dict] = []
    position_rows: list[dict] = []
    calendar_rows: list[dict] = []
    issue_rows: list[dict] = []
    tolerance = 0.02

    for week, event in enumerate(events, start=1):
        signal_date = pd.Timestamp(event["date"])
        execution_date = _next_nse_session(signal_date, holidays)
        target = {str(ticker): float(weight) for ticker, weight in event.get(weight_key, {}).items()}
        pre_trade = {str(ticker): float(weight) for ticker, weight in event.get(pre_weight_key, {}).items()}
        entries = _reason_map(event.get("entries"))
        exits = _reason_map(event.get("exits"))
        ranking = {ticker: rank for rank, ticker in enumerate(event.get("full_ranking", []), start=1)}
        all_tickers = sorted(set(pre_trade) | set(target))
        calculated_turnover = sum(abs(target.get(ticker, 0.0) - pre_trade.get(ticker, 0.0)) for ticker in all_tickers)
        engine_turnover = float(event.get(turnover_key, 0.0))
        turnover_difference = calculated_turnover - engine_turnover
        if abs(turnover_difference) < 1e-9:
            turnover_difference = 0.0
        target_sum = sum(target.values())
        pre_trade_sum = sum(pre_trade.values())
        weight_status = "OK" if abs(target_sum - 100.0) <= 0.02 else "FAIL"
        turnover_status = "OK" if abs(turnover_difference) <= tolerance else "FAIL"

        weekly_rows.append(
            {
                "Week": week,
                "Signal Date": signal_date,
                "Execution Date": execution_date,
                "Holdings": len(target),
                "Entries": len(entries),
                "Exits": len(exits),
                "Engine Turnover (%)": engine_turnover,
                "Calculated Turnover (%)": calculated_turnover,
                "Turnover Difference (pp)": turnover_difference,
                "Pre-trade Weight Sum (%)": pre_trade_sum,
                "Target Weight Sum (%)": target_sum,
                "Weight Check": weight_status,
                "Turnover Check": turnover_status,
                "Corporate Actions": len(event.get("corporate_actions", [])),
                "Valid Universe": int(event.get("valid_universe_size", 0)),
            }
        )
        calendar_rows.append(
            {
                "Week": week,
                "Signal Date": signal_date,
                "Expected Next NSE Session": execution_date,
                "Signal Weekday": signal_date.day_name(),
                "Execution Weekday": execution_date.strftime("%A"),
                "Holiday Shift": (execution_date - signal_date.date()).days > 1,
                "Status": "OK",
            }
        )

        if weight_status != "OK":
            issue_rows.append(
                {
                    "Severity": "HIGH",
                    "Category": "Weight tie-out",
                    "Signal Date": signal_date,
                    "Ticker": "",
                    "Observed": target_sum,
                    "Expected": 100.0,
                    "Difference": target_sum - 100.0,
                    "Explanation": "Post-rebalance target weights must sum to 100%.",
                }
            )
        if turnover_status != "OK":
            issue_rows.append(
                {
                    "Severity": "HIGH",
                    "Category": "Turnover reconciliation",
                    "Signal Date": signal_date,
                    "Ticker": "",
                    "Observed": calculated_turnover,
                    "Expected": engine_turnover,
                    "Difference": turnover_difference,
                    "Explanation": "Absolute pre-trade-to-target changes must match engine turnover.",
                }
            )

        for ticker in all_tickers:
            prior_weight = pre_trade.get(ticker, 0.0)
            target_weight = target.get(ticker, 0.0)
            delta = target_weight - prior_weight
            if abs(delta) <= 0.001:
                delta = 0.0
            action = "BUY" if delta > 0.001 else "SELL" if delta < -0.001 else "HOLD"
            position_rows.append(
                {
                    "Week": week,
                    "Signal Date": signal_date,
                    "Execution Date": execution_date,
                    "Ticker": ticker,
                    "Rank": ranking.get(ticker),
                    "Pre-trade Weight (%)": prior_weight,
                    "Target Weight (%)": target_weight,
                    "Trade Weight (%)": delta,
                    "Action": action,
                    "Entry Reason": entries.get(ticker, ""),
                    "Exit Reason": exits.get(ticker, ""),
                    "Latest Price Date": (
                        result.get("latest_price_dates", {}).get(ticker, "") if week == len(events) else ""
                    ),
                    "Tradability": (
                        result.get("tradability_status", {}).get(ticker, "") if week == len(events) else ""
                    ),
                }
            )

    freshness = result.get("data_freshness", {}) or {}
    provenance_rows = [
        {"Field": "Replay ID", "Value": result.get("strategy_fingerprint", "")},
        {"Field": "OHLCV source", "Value": result.get("ohlcv_source", "")},
        {"Field": "OHLCV as-of date", "Value": result.get("ohlcv_date", "")},
        {"Field": "Target session", "Value": freshness.get("target_date", "")},
        {"Field": "Verified common coverage", "Value": freshness.get("actual_latest_date", "")},
        {"Field": "Latest available price", "Value": freshness.get("max_price_date", "")},
        {"Field": "Refresh status", "Value": freshness.get("refresh_status", "")},
        {"Field": "Missing target symbols", "Value": freshness.get("missing_target_symbols", [])},
        {"Field": "Stale symbols", "Value": freshness.get("stale_symbols", [])},
        {"Field": "Refresh detail", "Value": freshness.get("error", "")},
    ]

    configuration_rows = [
        {"Parameter": "Portfolio source", "Value": params.get("portfolio_source", "snapshot")},
        {"Parameter": "Signal date", "Value": params.get("signal_date")},
        {"Parameter": "Portfolio start date", "Value": params.get("portfolio_start")},
        {"Parameter": "Band rule", "Value": params.get("band")},
        {"Parameter": "Variant", "Value": params.get("variant")},
        {"Parameter": "M (entry)", "Value": params.get("m")},
        {"Parameter": "N (exit)", "Value": params.get("n")},
        {"Parameter": "Ranking", "Value": params.get("sort_method")},
        {"Parameter": "Rebalance frequency", "Value": params.get("freq")},
        {"Parameter": "Index universe", "Value": params.get("indices", [])},
        {"Parameter": "Stage 2 entry filter", "Value": params.get("s2_entry")},
        {"Parameter": "Stage 2 entry threshold", "Value": params.get("s2_entry_threshold")},
        {"Parameter": "Stage 2 drop exit", "Value": params.get("s2_drop")},
        {"Parameter": "Stage 2 drop threshold", "Value": params.get("s2_threshold")},
        {"Parameter": "Max position (%)", "Value": params.get("max_pos")},
        {"Parameter": "Minimum history (sessions)", "Value": params.get("min_history")},
        {"Parameter": "Minimum annual return (%)", "Value": params.get("min_annual_return")},
        {"Parameter": "Within 52-week high (%)", "Value": params.get("pct_from_52w_high")},
        {"Parameter": "Maximum circuits (1 year)", "Value": params.get("max_circuits")},
        {"Parameter": "Close above 100 DMA", "Value": params.get("close_above_100dma")},
        {"Parameter": "Close above 200 DMA", "Value": params.get("close_above_200dma")},
        {"Parameter": "Positive days 3M (%)", "Value": params.get("pos_days_3m_min")},
        {"Parameter": "Positive days 6M (%)", "Value": params.get("pos_days_6m_min")},
        {"Parameter": "Positive days 12M (%)", "Value": params.get("pos_days_12m_min")},
        {"Parameter": "Transaction cost (%)", "Value": 0.1},
        {"Parameter": "STCG rate (%)", "Value": 20.0},
        {"Parameter": "LTCG rate (%)", "Value": 12.5},
    ]

    reconciliation_rows = list((reconciliation or {}).get("rows", []))
    for ticker in result.get("blocking_stale_incumbents", []):
        issue_rows.append(
            {
                "Severity": "HIGH",
                "Category": "Stale incumbent price",
                "Signal Date": pd.Timestamp(events[-1]["date"]) if events else pd.NaT,
                "Ticker": ticker,
                "Observed": result.get("tradability_status", {}).get(ticker, "NO DATA"),
                "Expected": "TRADABLE",
                "Difference": "",
                "Explanation": "A stale incumbent without a registered corporate action blocks execution.",
            }
        )
    if reconciliation:
        current_signal_date = pd.Timestamp(events[-1]["date"]) if events else pd.NaT
        for row in reconciliation_rows:
            ticker = str(row.get("Ticker", ""))
            if row.get("Action") == "HOLD":
                continue
            tradability = result.get("tradability_status", {}).get(ticker, "NO DATA")
            if tradability != "TRADABLE":
                issue_rows.append(
                    {
                        "Severity": "HIGH",
                        "Category": "Trade price unavailable",
                        "Signal Date": current_signal_date,
                        "Ticker": ticker,
                        "Observed": tradability,
                        "Expected": "TRADABLE",
                        "Difference": "",
                        "Explanation": "An executable order requires a current tradable price.",
                    }
                )
    replay_rows = replay_checks or []
    failed_replays = sum(str(row.get("Status", row.get("status", ""))).upper() == "FAIL" for row in replay_rows)
    errored_replays = sum(str(row.get("Status", row.get("status", ""))).upper() == "ERROR" for row in replay_rows)
    verification_status = (
        "NOT RUN" if replay_checks is None else "OK" if not failed_replays and not errored_replays else "FAIL"
    )
    overall_status = "OK" if not issue_rows and verification_status != "FAIL" else "FAIL"
    source = params.get("portfolio_source", "snapshot")
    scope_note = (
        "Model path plus current broker-snapshot reconciliation. Historical actual holdings are not asserted."
        if source == "snapshot"
        else "Model portfolio reconstructed from the selected portfolio start date."
    )
    summary_rows = [
        {"Metric": "Overall audit status", "Value": overall_status, "Status": overall_status, "Notes": ""},
        {"Metric": "Audit scope", "Value": source, "Status": "INFO", "Notes": scope_note},
        {"Metric": "Rebalance events", "Value": len(events), "Status": "INFO", "Notes": ""},
        {"Metric": "Position transitions", "Value": len(position_rows), "Status": "INFO", "Notes": ""},
        {"Metric": "Audit issues", "Value": len(issue_rows), "Status": overall_status, "Notes": ""},
        {
            "Metric": "Independent replay verification",
            "Value": verification_status,
            "Status": verification_status,
            "Notes": "Run the optional verification in LiveSignal for point-in-time reproducibility checks.",
        },
        {
            "Metric": "Generated at",
            "Value": pd.Timestamp.now(tz="Asia/Kolkata").tz_localize(None),
            "Status": "INFO",
            "Notes": "IST",
        },
        {"Metric": "Replay ID", "Value": result.get("strategy_fingerprint", ""), "Status": "INFO", "Notes": ""},
    ]

    if reconciliation:
        summary_rows.extend(
            [
                {
                    "Metric": "Portfolio value (₹)",
                    "Value": reconciliation.get("gross_value"),
                    "Status": "INFO",
                    "Notes": "Signal-date mark-to-market value.",
                },
                {
                    "Metric": "Actual trade turnover (%)",
                    "Value": reconciliation.get("turnover_pct"),
                    "Status": "INFO",
                    "Notes": "Whole-share broker reconciliation.",
                },
                {
                    "Metric": "Projected residual cash (₹)",
                    "Value": reconciliation.get("projected_cash"),
                    "Status": "INFO",
                    "Notes": "After the generated sell and buy quantities.",
                },
            ]
        )

    issue_columns = [
        "Severity",
        "Category",
        "Signal Date",
        "Ticker",
        "Observed",
        "Expected",
        "Difference",
        "Explanation",
    ]
    return {
        "Audit Summary": pd.DataFrame(summary_rows),
        "Configuration": pd.DataFrame(configuration_rows),
        "Data Provenance": pd.DataFrame(provenance_rows),
        "Weekly Rebalances": pd.DataFrame(weekly_rows),
        "Position Transitions": pd.DataFrame(position_rows),
        "Calendar Checks": pd.DataFrame(calendar_rows),
        "Current Reconciliation": (
            pd.DataFrame(reconciliation_rows)
            if reconciliation_rows
            else pd.DataFrame(
                [{"Status": "NOT APPLICABLE", "Detail": "No current broker reconciliation was supplied."}]
            )
        ),
        "Issues": pd.DataFrame(issue_rows, columns=issue_columns),
        "Replay Checks": (
            pd.DataFrame(replay_rows)
            if replay_rows
            else pd.DataFrame(
                [
                    {
                        "Status": "NOT RUN",
                        "Detail": "Independent point-in-time replay verification was not requested.",
                    }
                ]
            )
        ),
    }


def _number_format(header: str) -> str | None:
    lowered = header.lower()
    if "date" in lowered or "session" in lowered or "generated at" in lowered:
        return "yyyy-mm-dd"
    if "₹" in header:
        return "₹#,##0;[Red](₹#,##0);-"
    if "(%)" in header or "(pp)" in header or "turnover" in lowered or "weight" in lowered:
        return "0.00;[Red](0.00);-"
    if any(token in lowered for token in ("quantity", "holdings", "entries", "exits", "week", "rank")):
        return "#,##0;[Red](#,##0);-"
    return None


def _format_configuration_sheet(sheet) -> None:
    """Apply row-specific Excel types and formats to the heterogeneous Value column."""
    for row in range(2, sheet.max_row + 1):
        parameter = str(sheet.cell(row, 1).value or "")
        value_cell = sheet.cell(row, 2)
        if parameter in _CONFIGURATION_DATE_PARAMETERS:
            value_cell.number_format = "yyyy-mm-dd"
        elif parameter in _CONFIGURATION_INTEGER_PARAMETERS:
            value_cell.number_format = "#,##0"
        elif parameter in _CONFIGURATION_PERCENT_PARAMETERS:
            if isinstance(value_cell.value, (int, float)) and not isinstance(value_cell.value, bool):
                value_cell.value = float(value_cell.value) / 100.0
            value_cell.number_format = "0.##%"
        else:
            value_cell.number_format = "@"


def _write_table_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    columns = list(frame.columns) or ["Status"]
    sheet.append(columns)
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_serialise(value) for value in row])

    header = sheet[1]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(sheet.max_row, 1)}"

    if sheet.max_row > 1:
        table_name = "Audit" + "".join(ch for ch in sheet_name if ch.isalnum())
        table = Table(displayName=table_name[:250], ref=f"A1:{get_column_letter(len(columns))}{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for col_idx, column in enumerate(columns, start=1):
        fmt = _number_format(str(column))
        width = min(42, max(11, len(str(column)) + 2))
        for cell in sheet[get_column_letter(col_idx)]:
            if cell.row > 1:
                cell.alignment = Alignment(
                    vertical="top", wrap_text=isinstance(cell.value, str) and len(cell.value) > 35
                )
                if fmt:
                    cell.number_format = fmt
                width = min(42, max(width, min(len(str(cell.value or "")) + 2, 42)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    status_columns = [
        idx
        for idx, value in enumerate(columns, start=1)
        if "status" in str(value).lower() or "check" in str(value).lower()
    ]
    for col_idx in status_columns:
        letter = get_column_letter(col_idx)
        if sheet.max_row > 1:
            target = f"{letter}2:{letter}{sheet.max_row}"
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=_GREEN)),
            )
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=_RED)),
            )
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="equal", formula=['"NOT RUN"'], fill=PatternFill("solid", fgColor=_AMBER)),
            )


def build_live_signal_audit_workbook(
    params: dict,
    result: dict,
    reconciliation: dict | None = None,
    replay_checks: list[dict] | None = None,
) -> bytes:
    """Return a styled Excel audit workbook as bytes."""
    tables = build_live_signal_audit_tables(params, result, reconciliation, replay_checks)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, frame in tables.items():
        _write_table_sheet(workbook, sheet_name, frame)

    summary = workbook["Audit Summary"]
    summary.freeze_panes = None
    summary.insert_rows(1, amount=3)
    summary.merge_cells("A1:D1")
    summary["A1"] = "LiveSignal Portfolio Continuity Audit"
    summary["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    summary["A1"].font = Font(color=_WHITE, bold=True, size=16)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 28
    summary.merge_cells("A2:D2")
    summary["A2"] = "Point-in-time rebalance, transition, execution-calendar and broker-reconciliation evidence"
    summary["A2"].font = Font(color="475569", italic=True)
    summary["A2"].alignment = Alignment(wrap_text=True)
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 72
    for row in range(5, summary.max_row + 1):
        summary.cell(row, 1).font = Font(bold=True)
        summary.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
        metric = str(summary.cell(row, 1).value or "")
        if "Generated at" in metric:
            summary.cell(row, 2).number_format = "yyyy-mm-dd hh:mm"
        elif "(₹)" in metric:
            summary.cell(row, 2).number_format = "₹#,##0;[Red](₹#,##0);-"
        elif "(%)" in metric:
            summary.cell(row, 2).number_format = "0.00;[Red](0.00);-"
        status = str(summary.cell(row, 3).value or "")
        fill, color = (
            (_GREEN, _GREEN_TEXT)
            if status == "OK"
            else (_RED, _RED_TEXT) if status == "FAIL" else (_AMBER, _AMBER_TEXT)
        )
        summary.cell(row, 3).fill = PatternFill("solid", fgColor=fill)
        summary.cell(row, 3).font = Font(color=color, bold=True)
    for cell in summary[4]:
        cell.border = Border(bottom=Side(style="medium", color=_NAVY))
    summary.auto_filter.ref = f"A4:D{summary.max_row}"
    for table in summary.tables.values():
        table.ref = f"A4:D{summary.max_row}"

    _format_configuration_sheet(workbook["Configuration"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def audit_tables_as_json(tables: dict[str, pd.DataFrame]) -> dict[str, list[dict]]:
    """Return JSON-safe records for the CLI and programmatic consumers."""
    payload: dict[str, list[dict]] = {}
    for name, frame in tables.items():
        payload[name] = [
            {str(key): _serialise(value) for key, value in row.items()} for row in frame.to_dict("records")
        ]
    return payload
