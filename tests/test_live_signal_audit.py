import io
from datetime import date, datetime

from openpyxl import load_workbook

import app_live_signal as live
from live_signal_audit import build_live_signal_audit_tables, build_live_signal_audit_workbook


def _params() -> dict:
    return {
        "portfolio_source": "snapshot",
        "signal_date": date(2026, 9, 1),
        "portfolio_start": date(2025, 7, 2),
        "band": "classic",
        "variant": "Marginal Rebalance",
        "m": 2,
        "n": 3,
        "sort_method": "3 months",
        "freq": "weekly",
        "indices": ["Nifty 50"],
        "s2_drop": True,
        "s2_threshold": 3,
        "s2_entry": False,
        "s2_entry_threshold": 2,
        "max_pos": 0,
        "min_history": 252,
    }


def _result() -> dict:
    return {
        "portfolio_reset_date": date(2026, 8, 25),
        "strategy_fingerprint": "abc123",
        "ohlcv_date": "2026-09-01",
        "ohlcv_source": "parquet",
        "latest_price_dates": {"A": "2026-09-01", "C": "2026-09-01"},
        "tradability_status": {"A": "TRADABLE", "B": "TRADABLE", "C": "TRADABLE"},
        "data_freshness": {"target_date": "2026-09-01", "actual_latest_date": "2026-09-01"},
        "holdings_log": [
            {
                "date": date(2026, 8, 25),
                "holdings": ["A", "B"],
                "entries": ["A (Top-M)", "B (Top-M)"],
                "exits": [],
                "marg_weights": {"A": 50.0, "B": 50.0},
                "pre_rebalance_marg_weights": {},
                "marg_turnover_pct": 100.0,
                "full_ranking": ["A", "B", "C"],
                "valid_universe_size": 3,
                "corporate_actions": [],
            },
            {
                "date": date(2026, 9, 1),
                "holdings": ["A", "C"],
                "entries": ["C (Top-M)"],
                "exits": ["B (WRH)"],
                "marg_weights": {"A": 45.0, "C": 55.0},
                "pre_rebalance_marg_weights": {"A": 55.0, "B": 45.0},
                "marg_turnover_pct": 110.0,
                "full_ranking": ["C", "A", "B"],
                "valid_universe_size": 3,
                "corporate_actions": [],
            },
        ],
    }


def _reconciliation() -> dict:
    return {
        "gross_value": 1_000.0,
        "securities_value": 900.0,
        "cash": 100.0,
        "reserve_cash": 0.0,
        "projected_cash": 50.0,
        "turnover_pct": 55.0,
        "rows": [
            {
                "Ticker": "A",
                "Price": 100.0,
                "Actual quantity": 5,
                "Actual value (₹)": 500.0,
                "Actual weight (%)": 50.0,
                "Strategy target (%)": 45.0,
                "Target value (₹)": 450.0,
                "Target quantity": 4,
                "Order quantity": -1,
                "Action": "SELL",
                "Trade value (₹)": 100.0,
                "Projected value (₹)": 400.0,
                "Projected weight (%)": 40.0,
            }
        ],
    }


def test_audit_tables_cover_history_reconciliation_and_scope(monkeypatch):
    monkeypatch.setattr("live_signal_audit.load_nse_holidays", lambda: frozenset())

    tables = build_live_signal_audit_tables(_params(), _result(), _reconciliation())

    assert len(tables["Weekly Rebalances"]) == 2
    assert len(tables["Position Transitions"]) == 5
    assert len(tables["Current Reconciliation"]) == 1
    assert tables["Issues"].empty
    summary = tables["Audit Summary"].set_index("Metric")
    assert summary.loc["Overall audit status", "Value"] == "OK"
    assert summary.loc["Independent replay verification", "Value"] == "NOT RUN"
    assert "Historical actual holdings are not asserted" in summary.loc["Audit scope", "Notes"]


def test_audit_workbook_contains_all_expected_sheets(monkeypatch):
    monkeypatch.setattr("live_signal_audit.load_nse_holidays", lambda: frozenset())

    payload = build_live_signal_audit_workbook(_params(), _result(), _reconciliation())
    workbook = load_workbook(io.BytesIO(payload), data_only=False)

    assert payload[:2] == b"PK"
    assert workbook.sheetnames == [
        "Audit Summary",
        "Configuration",
        "Data Provenance",
        "Weekly Rebalances",
        "Position Transitions",
        "Calendar Checks",
        "Current Reconciliation",
        "Issues",
        "Replay Checks",
    ]
    assert workbook["Audit Summary"]["A1"].value == "LiveSignal Portfolio Continuity Audit"
    assert workbook["Weekly Rebalances"].max_row == 3


def test_configuration_sheet_uses_semantic_excel_types(monkeypatch):
    monkeypatch.setattr("live_signal_audit.load_nse_holidays", lambda: frozenset())
    params = _params()
    params["max_pos"] = 15

    payload = build_live_signal_audit_workbook(params, _result(), _reconciliation())
    sheet = load_workbook(io.BytesIO(payload), data_only=False)["Configuration"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2) for row in range(2, sheet.max_row + 1)}

    assert isinstance(values["Signal date"].value, datetime)
    assert values["Signal date"].number_format == "yyyy-mm-dd"
    assert values["M (entry)"].value == 2
    assert values["M (entry)"].data_type == "n"
    assert values["M (entry)"].number_format == "#,##0"
    assert values["Stage 2 entry filter"].value == "No"
    assert values["Stage 2 entry filter"].number_format == "@"
    assert values["Max position (%)"].value == 0.15
    assert values["Max position (%)"].number_format == "0.##%"
    assert values["Transaction cost (%)"].value == 0.001
    assert values["STCG rate (%)"].value == 0.2
    assert values["LTCG rate (%)"].value == 0.125


def test_audit_flags_turnover_mismatch(monkeypatch):
    monkeypatch.setattr("live_signal_audit.load_nse_holidays", lambda: frozenset())
    result = _result()
    result["holdings_log"][-1]["marg_turnover_pct"] = 10.0

    tables = build_live_signal_audit_tables(_params(), result)

    assert "Turnover reconciliation" in set(tables["Issues"]["Category"])
    summary = tables["Audit Summary"].set_index("Metric")
    assert summary.loc["Overall audit status", "Value"] == "FAIL"


def test_independent_replay_checks_selected_variant(monkeypatch):
    result = _result()

    def fake_run(params):
        signal = params["signal_date"]
        expected = next(event for event in result["holdings_log"] if event["date"] == signal)
        return {"holdings_log": [expected]}

    monkeypatch.setattr(live, "_run_signal", fake_run)

    checks = live._run_independent_replay_checks(_params(), result)

    assert len(checks) == 2
    assert {row["Status"] for row in checks} == {"OK"}


def test_audit_method_comparison_explains_both_levels():
    comparison = live._audit_method_comparison().set_index("Aspect")

    assert "historical replay already calculated" in comparison.loc["Source", "Standard Audit Workbook"]
    assert "Recalculates every rebalance date" in comparison.loc["Source", "Independent Weekly Replay Verification"]
    assert "hidden state leakage" in comparison.loc["Detects", "Independent Weekly Replay Verification"]
