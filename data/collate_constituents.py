"""
collate_constituents.py
-----------------------
Collates three NSE index constituent data files into a single unified CSV.

Source files (all in the same directory as this script):
  1. index_const.csv                                  – 2020-03-31 to 2024-03-28
  2. IndexInclExcl - Upto 2020.xls                   – historical, pre-2020
  3. N750 Historical Constituents - 202409 onwards.xlsx – 2024-09-30 onwards

Output:
  unified_constituents.csv  (same directory)
  Schema: INDEX_NAME, TIME_STAMP, SYMBOL, INDUSTRY, CAP_WEIGHT
"""

# ---------------------------------------------------------------------------
# 0. Dependency bootstrap
# ---------------------------------------------------------------------------
import importlib
import importlib.util
import subprocess
import sys


def _ensure(pkg, import_name=None):
    name = import_name or pkg
    if importlib.util.find_spec(name) is None:
        print(f"[INFO] Installing {pkg} ...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
        except subprocess.CalledProcessError:
            print(f"[WARN] Could not auto-install {pkg}. " f"Please run:  pip install {pkg}")


_ensure("xlrd")
_ensure("openpyxl")

# ---------------------------------------------------------------------------
# 1. Imports
# ---------------------------------------------------------------------------
import warnings
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# 2. Constants
# ---------------------------------------------------------------------------
DATA_DIR = Path(__file__).parent

CSV_PATH = DATA_DIR / "index_const.csv"
XLS_PATH = DATA_DIR / "IndexInclExcl - Upto 2020.xls"
XLSX_PATH = DATA_DIR / "N750 Historical Constituents - 202409 onwards.xlsx"
OUTPUT_PATH = DATA_DIR / "unified_constituents.csv"

SCHEMA = ["INDEX_NAME", "TIME_STAMP", "SYMBOL", "INDUSTRY", "CAP_WEIGHT"]
DEDUP_KEYS = ["INDEX_NAME", "TIME_STAMP", "SYMBOL"]

# xlsx sheet-name -> NSE index name (Sheet8 is a note; N100 is empty — both excluded)
SHEET_TO_INDEX = {
    "N50": "NIFTY 50",
    "NN50": "NIFTY NEXT 50",
    "N100": "NIFTY 100",
    "N200": "NIFTY 200",
    "MC150": "NIFTY MIDCAP 150",
    "SC250": "NIFTY SMALLCAP 250",
    "Micro250": "NIFTY MICROCAP 250",
}

# Heuristic fragments for xls column name normalisation (lowercase match)
COLUMN_HEURISTICS = {
    "INDEX_NAME": ["index_name", "indexname", "index name", "index"],
    "TIME_STAMP": ["time_stamp", "timestamp", "date", "time", "month", "period", "effective"],
    "SYMBOL": ["symbol", "ticker", "scrip", "stock", "nse_code", "nse code"],
    "INDUSTRY": ["industry", "sector", "group"],
    "CAP_WEIGHT": ["cap_weight", "capwt", "weight", "wt", "cap weight"],
}

# ---------------------------------------------------------------------------
# 3. Helper functions
# ---------------------------------------------------------------------------


def _excel_serial_to_date(value) -> str | None:
    """Convert an Excel serial date number (or Python datetime) to YYYY-MM-DD."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.strftime("%Y-%m-%d")
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(float(value)))).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OverflowError):
        return None


def _normalize_xls_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Map arbitrary xls column names to the unified schema using heuristics."""
    # Normalise column names for matching
    col_clean = {col: col.strip().lower().replace(" ", "_").replace("-", "_") for col in df.columns}

    col_map = {}  # original_col -> SCHEMA_col
    used_targets: set[str] = set()

    for orig_col, clean in col_clean.items():
        for target, fragments in COLUMN_HEURISTICS.items():
            if target in used_targets:
                continue
            if any(frag in clean for frag in fragments):
                col_map[orig_col] = target
                used_targets.add(target)
                break

    df = df.rename(columns=col_map)

    # Fill any missing schema columns with NA
    for col in SCHEMA:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[SCHEMA].copy()

    # Normalise TIME_STAMP — try dayfirst=True for DD-MM-YYYY formats common in Indian data
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        df["TIME_STAMP"] = pd.to_datetime(df["TIME_STAMP"], errors="coerce", dayfirst=True).dt.strftime("%Y-%m-%d")

    # If INDEX_NAME is entirely missing, fall back to sheet name
    if df["INDEX_NAME"].isna().all():
        df["INDEX_NAME"] = sheet_name

    # Drop rows where the critical identifying fields are absent
    df = df.dropna(subset=["SYMBOL", "TIME_STAMP"])
    return df


# ---------------------------------------------------------------------------
# 4. Parser: File 1 — index_const.csv
# ---------------------------------------------------------------------------


def parse_csv() -> pd.DataFrame:
    print(f"[1/3] Reading {CSV_PATH.name} ...")

    df = pd.read_csv(
        CSV_PATH,
        skiprows=[1],  # row index 1 = dashes separator line
        on_bad_lines="skip",
        dtype=str,
        encoding="utf-8",
        encoding_errors="replace",
    )

    # Standardise column names
    df.columns = [c.strip() for c in df.columns]

    # Drop SQL rowcount footer  e.g. "(119707 rows affected)"
    df = df[~df["INDEX_NAME"].str.startswith("(", na=False)].copy()

    # Normalise types
    df["TIME_STAMP"] = pd.to_datetime(df["TIME_STAMP"], errors="coerce").dt.strftime("%Y-%m-%d")
    df["CAP_WEIGHT"] = pd.to_numeric(df["CAP_WEIGHT"], errors="coerce")
    df["SYMBOL"] = df["SYMBOL"].str.strip().str.upper()
    df["INDEX_NAME"] = df["INDEX_NAME"].str.strip()

    df = df[SCHEMA].dropna(subset=["SYMBOL", "TIME_STAMP"])
    print(f"       -> {len(df):,} rows loaded")
    return df


# ---------------------------------------------------------------------------
# 5. Parser: File 2 — IndexInclExcl - Upto 2020.xls
# ---------------------------------------------------------------------------


def parse_xls() -> pd.DataFrame:
    print(f"[2/3] Reading {XLS_PATH.name} ...")

    if not XLS_PATH.exists():
        print("       [WARN] File not found — skipping.")
        return pd.DataFrame(columns=SCHEMA)

    try:
        sheets: dict[str, pd.DataFrame] = pd.read_excel(
            XLS_PATH,
            sheet_name=None,
            engine="xlrd",
            dtype=str,
        )
    except Exception as exc:
        print(f"       [WARN] Could not read .xls file ({exc}) — skipping.")
        return pd.DataFrame(columns=SCHEMA)

    frames = []
    for sheet_name, sheet_df in sheets.items():
        if sheet_df.empty:
            continue
        try:
            normalized = _normalize_xls_sheet(sheet_df, sheet_name)
            frames.append(normalized)
            print(f"       Sheet '{sheet_name}': {len(normalized):,} rows")
        except Exception as exc:
            print(f"       [WARN] Sheet '{sheet_name}' failed ({exc}) — skipping sheet.")

    if not frames:
        print("       -> 0 rows loaded (no valid sheets)")
        return pd.DataFrame(columns=SCHEMA)

    df = pd.concat(frames, ignore_index=True)
    df["SYMBOL"] = df["SYMBOL"].str.strip().str.upper()
    df["INDEX_NAME"] = df["INDEX_NAME"].str.strip()
    print(f"       -> {len(df):,} rows loaded")
    return df


# ---------------------------------------------------------------------------
# 6. Parser: File 3 — N750 Historical Constituents - 202409 onwards.xlsx
# ---------------------------------------------------------------------------


def parse_xlsx() -> pd.DataFrame:
    print(f"[3/3] Reading {XLSX_PATH.name} ...")

    if not XLSX_PATH.exists():
        print("       [WARN] File not found — skipping.")
        return pd.DataFrame(columns=SCHEMA)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("       [WARN] openpyxl not available — skipping.")
        return pd.DataFrame(columns=SCHEMA)

    records = []

    try:
        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    except Exception as exc:
        print(f"       [WARN] Could not open .xlsx file ({exc}) — skipping.")
        return pd.DataFrame(columns=SCHEMA)

    for sheet_tab, index_name in SHEET_TO_INDEX.items():
        if sheet_tab not in wb.sheetnames:
            print(f"       [WARN] Sheet '{sheet_tab}' not found — skipping.")
            continue

        ws = wb[sheet_tab]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            print(f"       Sheet '{sheet_tab}': empty — skipping.")
            continue

        # Row 0: date serial numbers (or datetime objects)
        header_row = rows[0]
        col_dates: dict[int, str] = {}
        for col_idx, val in enumerate(header_row):
            if val is None:
                continue
            date_str = _excel_serial_to_date(val)
            if date_str:
                col_dates[col_idx] = date_str

        if not col_dates:
            print(f"       Sheet '{sheet_tab}': no valid date headers — skipping.")
            continue

        # Rows 1+: symbol grid
        sheet_records = 0
        for row in rows[1:]:
            for col_idx, date_str in col_dates.items():
                if col_idx >= len(row):
                    continue
                symbol = row[col_idx]
                if symbol is None or str(symbol).strip() == "":
                    continue
                records.append(
                    {
                        "INDEX_NAME": index_name,
                        "TIME_STAMP": date_str,
                        "SYMBOL": str(symbol).strip().upper(),
                        "INDUSTRY": pd.NA,
                        "CAP_WEIGHT": pd.NA,
                    }
                )
                sheet_records += 1

        print(f"       Sheet '{sheet_tab}' -> '{index_name}': {sheet_records:,} rows")

    wb.close()

    if not records:
        print("       -> 0 rows loaded")
        return pd.DataFrame(columns=SCHEMA)

    df = pd.DataFrame(records, columns=SCHEMA)
    print(f"       -> {len(df):,} rows loaded total")
    return df


# ---------------------------------------------------------------------------
# 7. Main: union -> dedup -> sort -> write -> verify
# ---------------------------------------------------------------------------


def main():
    print("=" * 60)
    print("  NSE Constituent Collation Script")
    print("=" * 60)

    df1 = parse_csv()
    df2 = parse_xls()
    df3 = parse_xlsx()

    print("\n[Collating] Concatenating all sources ...")
    # File 1 first -> its rows win deduplication (richest data)
    combined_before_dedup = pd.concat([df1, df2, df3], ignore_index=True)

    # Normalise key fields before dedup
    combined_before_dedup["SYMBOL"] = combined_before_dedup["SYMBOL"].str.strip().str.upper()
    combined_before_dedup["INDEX_NAME"] = combined_before_dedup["INDEX_NAME"].str.strip()

    rows_before = len(combined_before_dedup)
    print(f"         Rows before dedup: {rows_before:,}")

    combined = combined_before_dedup.drop_duplicates(subset=DEDUP_KEYS, keep="first")
    rows_after = len(combined)
    print(f"         Rows after dedup:  {rows_after:,}  (removed {rows_before - rows_after:,} duplicates)")

    print("\n[Sorting] INDEX_NAME -> TIME_STAMP -> SYMBOL ...")
    combined = combined.sort_values(DEDUP_KEYS, ascending=True).reset_index(drop=True)

    print(f"\n[Writing] {OUTPUT_PATH} ...")
    combined.to_csv(
        OUTPUT_PATH,
        index=False,
        encoding="utf-8",
        float_format="%.15g",  # avoids scientific notation for small CAP_WEIGHT values
    )
    print(f"         Done — {rows_after:,} rows written.")

    # -----------------------------------------------------------------------
    # Verification
    # -----------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("  VERIFICATION")
    print("=" * 60)

    verify = pd.read_csv(OUTPUT_PATH, dtype=str)

    print(f"  Total rows       : {len(verify):,}")
    print(f"  Columns          : {verify.columns.tolist()}")
    print(f"  Unique INDEX_NAME: {verify['INDEX_NAME'].nunique()}")
    print(f"  Unique SYMBOL    : {verify['SYMBOL'].nunique():,}")

    ts_valid = pd.to_datetime(verify["TIME_STAMP"], errors="coerce").dropna()
    if not ts_valid.empty:
        print(f"  TIME_STAMP range : {ts_valid.min().date()} -> {ts_valid.max().date()}")

    dupes = verify.duplicated(subset=DEDUP_KEYS).sum()
    print(f"  Duplicate keys   : {dupes}  {'OK OK' if dupes == 0 else '!! PROBLEM'}")

    missing_sym = verify["SYMBOL"].isna().sum()
    missing_idx = verify["INDEX_NAME"].isna().sum()
    missing_ts = verify["TIME_STAMP"].isna().sum()
    print(f"  Missing SYMBOL   : {missing_sym}  {'OK OK' if missing_sym == 0 else '!! PROBLEM'}")
    print(f"  Missing INDEX    : {missing_idx}  {'OK OK' if missing_idx == 0 else '!! PROBLEM'}")
    print(f"  Missing TIME_STAMP: {missing_ts}  {'OK OK' if missing_ts == 0 else '!! PROBLEM'}")

    # Spot check: NIFTY 50 at 2024-09-30 (from File 3)
    spot = verify[(verify["INDEX_NAME"] == "NIFTY 50") & (verify["TIME_STAMP"] == "2024-09-30")]
    print(f"\n  Spot check — NIFTY 50 @ 2024-09-30: {len(spot)} rows (expect ~50)")
    if not spot.empty:
        print(f"  First 5 symbols : {spot['SYMBOL'].head().tolist()}")

    # Source contribution summary
    print("\n  Source breakdown:")
    print(f"    File 1 (index_const.csv)                 : {len(df1):>7,} rows")
    print(f"    File 2 (IndexInclExcl - Upto 2020.xls)  : {len(df2):>7,} rows")
    print(f"    File 3 (N750 xlsx, 202409 onwards)       : {len(df3):>7,} rows")
    print(f"    Combined before dedup                    : {rows_before:>7,} rows")
    print(f"    After dedup (final output)               : {rows_after:>7,} rows")

    print("\n" + "=" * 60)
    print(f"  Output saved to: {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
