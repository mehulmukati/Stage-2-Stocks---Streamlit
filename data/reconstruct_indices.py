"""
reconstruct_indices.py
----------------------
Reconstructs NSE index compositions as a full time series by combining:
  - File 1: index_const.csv (quarterly snapshots, 2020-03-31 to 2024-03-28)
  - File 2: IndexInclExcl - Upto 2020.xls (event log, 1996-09-18 to 2020-11-02)
  - File 3: N750 Historical Constituents - 202409 onwards.xlsx (semi-annual snapshots)

Algorithm (per-index):
  1. Anchor on earliest F1 (or F3) snapshot for this index.
  2. Walk BACKWARD through F2 events to reconstruct pre-anchor history.
  3. Walk FORWARD applying F2 post-anchor events; snap to F1/F3 snapshots.
  4. Emit composition row for every event-date and snapshot-date.
  5. Log discrepancies to validation_report.csv.

Outputs (all in data/):
  - compositions.parquet     : main time-series (long form)
  - events.csv               : unified event log
  - scrip_symbol_map.csv     : name -> ticker mapping
  - scrip_symbol_unresolved.csv : low-confidence or failed matches
  - manual_overrides.json    : stub for human curations (wins all)
  - validation_report.csv    : discrepancies between reconstruction and snapshots
  - ../constituents.json     : refreshed with latest composition
"""

# =============================================================================
# 0. Bootstrap
# =============================================================================
import importlib
import importlib.util
import subprocess
import sys


def _ensure(pkg, import_name=None):
    name = import_name or pkg
    if importlib.util.find_spec(name) is None:
        print(f"[INFO] Installing {pkg} ...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
        except subprocess.CalledProcessError:
            print(f"[WARN] Could not auto-install {pkg}; run: pip install {pkg}")


_ensure("xlrd")
_ensure("openpyxl")
_ensure("rapidfuzz")

# =============================================================================
# 1. Imports
# =============================================================================
import csv
import json
import re
import warnings
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process

# =============================================================================
# 2. Constants
# =============================================================================
DATA_DIR = Path(__file__).parent
PROJECT_ROOT = DATA_DIR.parent

CSV_PATH = DATA_DIR / "index_const.csv"
XLS_PATH = DATA_DIR / "IndexInclExcl - Upto 2020.xls"
XLSX_PATH = DATA_DIR / "N750 Historical Constituents - 202409 onwards.xlsx"

OUT_COMPOSITIONS = DATA_DIR / "compositions.parquet"
OUT_EVENTS = DATA_DIR / "events.csv"
OUT_MAP = DATA_DIR / "scrip_symbol_map.csv"
OUT_UNRESOLVED = DATA_DIR / "scrip_symbol_unresolved.csv"
OUT_OVERRIDES = DATA_DIR / "manual_overrides.json"
OUT_VALIDATION = DATA_DIR / "validation_report.csv"
OUT_CONSTITUENTS = PROJECT_ROOT / "constituents.json"

# File 3 sheet name -> canonical index name (N100 empty, Sheet8 is a note)
SHEET_TO_INDEX = {
    "N50": "NIFTY 50",
    "NN50": "NIFTY NEXT 50",
    "N100": "NIFTY 100",
    "N200": "NIFTY 200",
    "MC150": "NIFTY MIDCAP 150",
    "SC250": "NIFTY SMALLCAP 250",
    "Micro250": "NIFTY MICROCAP 250",
}

# Canonicalization suffixes/tokens to strip during name-symbol matching
NAME_STRIPS = [
    r"\blimited\b",
    r"\bltd\.?",
    r"\bpvt\.?",
    r"\bprivate\b",
    r"\bcorporation\b",
    r"\bcorp\.?",
    r"\bincorporated\b",
    r"\binc\.?",
    r"\bcompany\b",
    r"\bco\.?",
    r"\band\b",
    r"\(india\)",
    r"\(the\)",
    r"\b&\b",
    r"\bindia\b",
    r"\s+",
]

# Fuzzy-match thresholds
THRESHOLD_HIGH = 90  # auto-accept
THRESHOLD_MED = 70  # flag as needs_review


# =============================================================================
# 3. Helpers
# =============================================================================
def log(msg, level="INFO"):
    print(f"[{level}] {msg}")


def canonicalize_name(s: str) -> str:
    """Strip suffixes, punctuation, spaces; uppercase. Used for name<->ticker matching."""
    if s is None:
        return ""
    s = str(s).lower().strip()
    # Remove non-breaking space + punctuation
    s = s.replace("\xa0", " ").replace(".", "").replace(",", "").replace("'", "")
    s = s.replace("(", " ").replace(")", " ").replace("-", " ").replace("/", " ")
    # Apply strip patterns
    for pat in NAME_STRIPS:
        s = re.sub(pat, " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", "", s)
    return s.upper()


def canonicalize_symbol(s: str) -> str:
    """Uppercase, strip hyphens/ampersands/spaces."""
    if s is None:
        return ""
    return str(s).upper().strip().replace("&", "").replace("-", "").replace(" ", "")


def canonicalize_index(s: str) -> str:
    """Normalize index names to one canonical form."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_description(s: str) -> str:
    """Map any Description variant to one of {INCL, EXCL}. Returns None if unknown."""
    if s is None:
        return None
    s = re.sub(r"\s+", " ", str(s).strip().lower())
    if "inclus" in s and "index" in s:
        return "INCL"
    if "exclus" in s and "index" in s:
        return "EXCL"
    return None


def excel_serial_to_date(value):
    """Convert Excel serial date number or datetime to YYYY-MM-DD string."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    try:
        return (datetime(1899, 12, 30) + timedelta(days=int(float(value)))).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OverflowError):
        return None


# =============================================================================
# 4. STAGE S1 — Parse all three files
# =============================================================================
def parse_f1():
    """Parse index_const.csv. Returns DataFrame with INDEX_NAME, TIME_STAMP, SYMBOL, INDUSTRY, CAP_WEIGHT."""
    log(f"S1.F1: reading {CSV_PATH.name}")
    # QUOTE_NONE is essential: the source CSV has unclosed quote marks in
    # INDUSTRY values (e.g., TITAN's "GEMS, JEWELLERY ...") that otherwise
    # cause the parser to swallow dozens of subsequent rows as one logical
    # multi-line field. Python engine handles malformed rows (the ~41 with
    # unquoted embedded commas) more gracefully than the C engine.
    df = pd.read_csv(
        CSV_PATH,
        skiprows=[1],
        on_bad_lines="skip",
        dtype=str,
        encoding="utf-8",
        encoding_errors="replace",
        engine="python",
        quoting=csv.QUOTE_NONE,
    )
    df.columns = [c.strip() for c in df.columns]
    df = df[~df["INDEX_NAME"].str.startswith("(", na=False)].copy()
    df["TIME_STAMP"] = pd.to_datetime(df["TIME_STAMP"], errors="coerce").dt.strftime("%Y-%m-%d")
    df["CAP_WEIGHT"] = pd.to_numeric(df["CAP_WEIGHT"], errors="coerce")
    df["SYMBOL"] = df["SYMBOL"].str.strip().str.upper()
    df["INDEX_NAME"] = df["INDEX_NAME"].map(canonicalize_index)
    df = df.dropna(subset=["SYMBOL", "TIME_STAMP", "INDEX_NAME"])
    df = df.drop_duplicates(subset=["INDEX_NAME", "TIME_STAMP", "SYMBOL"])
    log(
        f"       -> F1: {len(df):,} rows; {df['INDEX_NAME'].nunique()} indices; "
        f"{df['SYMBOL'].nunique():,} symbols; dates {df['TIME_STAMP'].min()} to {df['TIME_STAMP'].max()}"
    )
    return df


def parse_f2():
    """Parse IndexInclExcl xls. Returns DataFrame with INDEX_NAME, EVENT_DATE, SCRIP_NAME, EVENT_TYPE."""
    log(f"S1.F2: reading {XLS_PATH.name}")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        sheets = pd.read_excel(XLS_PATH, sheet_name=None, engine="xlrd", dtype=object)

    frames = []
    for sheet_name, sdf in sheets.items():
        if sdf.empty:
            continue
        sdf.columns = [str(c).strip() for c in sdf.columns]
        required = {"Index Name", "Event Date", "Scrip Name", "Description"}
        if not required.issubset(set(sdf.columns)):
            log(f"       [WARN] sheet '{sheet_name}' missing required cols; skipping", "WARN")
            continue
        sdf = sdf[["Index Name", "Event Date", "Scrip Name", "Description"]].copy()
        sdf.columns = ["INDEX_NAME", "EVENT_DATE", "SCRIP_NAME", "_DESC"]
        sdf["INDEX_NAME"] = sdf["INDEX_NAME"].map(canonicalize_index)
        sdf["SCRIP_NAME"] = sdf["SCRIP_NAME"].astype(str).str.strip()
        sdf["EVENT_TYPE"] = sdf["_DESC"].map(normalize_description)

        # Parse dates: mix of datetime objects and DD-MM-YYYY strings
        def _parse_dt(v):
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d")
            try:
                return pd.to_datetime(v, dayfirst=True, errors="coerce").strftime("%Y-%m-%d")
            except Exception:
                return None

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            sdf["EVENT_DATE"] = sdf["EVENT_DATE"].map(_parse_dt)
        sdf = sdf.drop(columns=["_DESC"])
        sdf = sdf.dropna(subset=["INDEX_NAME", "EVENT_DATE", "SCRIP_NAME", "EVENT_TYPE"])
        sdf = sdf[sdf["SCRIP_NAME"].str.len() > 0]
        frames.append(sdf)

    df = (
        pd.concat(frames, ignore_index=True)
        if frames
        else pd.DataFrame(columns=["INDEX_NAME", "EVENT_DATE", "SCRIP_NAME", "EVENT_TYPE"])
    )
    df = df.drop_duplicates(subset=["INDEX_NAME", "EVENT_DATE", "SCRIP_NAME", "EVENT_TYPE"])
    log(
        f"       -> F2: {len(df):,} events; {df['INDEX_NAME'].nunique()} indices; "
        f"{df['SCRIP_NAME'].nunique():,} scrip names; "
        f"INCL={len(df[df['EVENT_TYPE'] == 'INCL']):,}, EXCL={len(df[df['EVENT_TYPE'] == 'EXCL']):,}; "
        f"dates {df['EVENT_DATE'].min()} to {df['EVENT_DATE'].max()}"
    )
    return df


def parse_f3():
    """Parse N750 xlsx. Returns DataFrame with INDEX_NAME, TIME_STAMP, SYMBOL."""
    log(f"S1.F3: reading {XLSX_PATH.name}")
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    records = []
    for sheet_tab, index_name in SHEET_TO_INDEX.items():
        if sheet_tab not in wb.sheetnames:
            continue
        ws = wb[sheet_tab]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        col_dates = {}
        for ci, v in enumerate(header):
            if v is None:
                continue
            d = excel_serial_to_date(v)
            if d:
                col_dates[ci] = d
        if not col_dates:
            continue
        for row in rows[1:]:
            for ci, dstr in col_dates.items():
                if ci >= len(row):
                    continue
                sym = row[ci]
                if sym is None or str(sym).strip() == "":
                    continue
                records.append(
                    {
                        "INDEX_NAME": canonicalize_index(index_name),
                        "TIME_STAMP": dstr,
                        "SYMBOL": str(sym).strip().upper(),
                    }
                )
    wb.close()
    df = pd.DataFrame(records)
    if not df.empty:
        df = df.drop_duplicates(subset=["INDEX_NAME", "TIME_STAMP", "SYMBOL"])
    log(
        f"       -> F3: {len(df):,} rows; {df['INDEX_NAME'].nunique() if len(df) else 0} indices; "
        f"dates {df['TIME_STAMP'].min() if len(df) else 'NA'} to "
        f"{df['TIME_STAMP'].max() if len(df) else 'NA'}"
    )
    return df


# =============================================================================
# 5. STAGE S2 — Build SCRIP_NAME -> SYMBOL mapping
# =============================================================================
def load_manual_overrides() -> dict:
    """Load human-curated overrides. Wins all automatic matches."""
    if OUT_OVERRIDES.exists():
        try:
            raw = json.loads(OUT_OVERRIDES.read_text("utf-8"))
            # Skip metadata keys (start with _) and any non-string values (e.g. example dict)
            return {
                k.strip(): v.strip().upper() for k, v in raw.items() if not k.startswith("_") and isinstance(v, str)
            }
        except Exception as e:
            log(f"[WARN] could not parse {OUT_OVERRIDES.name}: {e}", "WARN")
    else:
        # Create empty stub so users know where to add overrides
        OUT_OVERRIDES.write_text(
            json.dumps(
                {
                    "_comment": "Hand-curated SCRIP_NAME -> SYMBOL overrides. "
                    "Edit this file to fix mapping errors. Wins all automatic matches.",
                    "_example": {"Some Company Ltd.": "SOMECO"},
                },
                indent=2,
            ),
            encoding="utf-8",
        )
    return {}


def build_mapping(f2_scrip_names, symbol_universe):
    """
    Build canonical scrip_name -> symbol mapping.
    Pipeline:
      1. Apply manual overrides (win all).
      2. Exact canonical-to-canonical match against symbol universe.
      3. rapidfuzz WRatio >= 90 -> high confidence auto-match.
      4. 70-89 -> flagged as medium.
      5. < 70 -> unresolved.
    """
    log(f"S2: building name-symbol map ({len(f2_scrip_names)} names vs {len(symbol_universe)} symbols)")

    overrides = load_manual_overrides()
    log(f"     manual overrides loaded: {len(overrides)}")

    # Build canonical form of symbols with originals as values
    sym_canonical = {}  # canonical -> original symbol
    for s in symbol_universe:
        canon = canonicalize_symbol(s)
        if canon and canon not in sym_canonical:
            sym_canonical[canon] = s
    canonical_list = list(sym_canonical.keys())

    mapping_rows = []
    for name in f2_scrip_names:
        name_clean = str(name).strip()
        if not name_clean:
            continue

        # 1. Manual override
        if name_clean in overrides:
            mapping_rows.append(
                {
                    "SCRIP_NAME": name_clean,
                    "SYMBOL": overrides[name_clean],
                    "METHOD": "manual_override",
                    "CONFIDENCE": "high",
                    "SCORE": 100.0,
                }
            )
            continue

        canon = canonicalize_name(name_clean)

        # 2. Exact canonical match
        if canon and canon in sym_canonical:
            mapping_rows.append(
                {
                    "SCRIP_NAME": name_clean,
                    "SYMBOL": sym_canonical[canon],
                    "METHOD": "exact_canonical",
                    "CONFIDENCE": "high",
                    "SCORE": 100.0,
                }
            )
            continue

        # 3. Fuzzy match via rapidfuzz (WRatio is robust to order/subsets)
        if canon and canonical_list:
            best = process.extractOne(canon, canonical_list, scorer=fuzz.WRatio, score_cutoff=THRESHOLD_MED)
            if best is not None:
                matched_canon, score, _ = best
                if score >= THRESHOLD_HIGH:
                    mapping_rows.append(
                        {
                            "SCRIP_NAME": name_clean,
                            "SYMBOL": sym_canonical[matched_canon],
                            "METHOD": "fuzzy_high",
                            "CONFIDENCE": "high",
                            "SCORE": float(score),
                        }
                    )
                    continue
                else:  # 70-89
                    mapping_rows.append(
                        {
                            "SCRIP_NAME": name_clean,
                            "SYMBOL": sym_canonical[matched_canon],
                            "METHOD": "fuzzy_med",
                            "CONFIDENCE": "medium",
                            "SCORE": float(score),
                        }
                    )
                    continue

        # 4. Unresolved
        mapping_rows.append(
            {
                "SCRIP_NAME": name_clean,
                "SYMBOL": None,
                "METHOD": "unresolved",
                "CONFIDENCE": "none",
                "SCORE": 0.0,
            }
        )

    df_map = pd.DataFrame(mapping_rows).drop_duplicates(subset=["SCRIP_NAME"])

    # Save outputs
    df_map.to_csv(OUT_MAP, index=False, encoding="utf-8")
    df_unresolved = df_map[df_map["METHOD"].isin(["fuzzy_med", "unresolved"])].sort_values("SCORE", ascending=False)
    df_unresolved.to_csv(OUT_UNRESOLVED, index=False, encoding="utf-8")

    breakdown = df_map["METHOD"].value_counts().to_dict()
    log(f"     mapping breakdown: {breakdown}")
    log(f"     written: {OUT_MAP.name} ({len(df_map)} rows), {OUT_UNRESOLVED.name} ({len(df_unresolved)} rows)")
    return df_map


# =============================================================================
# 6. STAGE S3 — Resolve events (join mapping, normalize, discard inception EXCLs)
# =============================================================================
def resolve_events(df_f2, df_map):
    """
    Join mapping onto F2 events; discard inception-day Exclusions per user rule.
    Returns events DataFrame with SYMBOL column.
    """
    log("S3: resolving events (joining mapping, inception rule)")
    name_to_sym = dict(zip(df_map["SCRIP_NAME"], df_map["SYMBOL"]))
    name_to_conf = dict(zip(df_map["SCRIP_NAME"], df_map["CONFIDENCE"]))
    name_to_score = dict(zip(df_map["SCRIP_NAME"], df_map["SCORE"]))

    df = df_f2.copy()
    df["SYMBOL"] = df["SCRIP_NAME"].map(name_to_sym)
    df["CONFIDENCE"] = df["SCRIP_NAME"].map(name_to_conf).fillna("none")
    df["MATCH_SCORE"] = df["SCRIP_NAME"].map(name_to_score).fillna(0.0)
    df["SOURCE"] = "F2"

    # Inception-day Exclusion rule: on each index's earliest event date, drop EXCLs
    earliest = df.groupby("INDEX_NAME")["EVENT_DATE"].min().to_dict()
    inception_excl_mask = df.apply(
        lambda r: r["EVENT_TYPE"] == "EXCL" and r["EVENT_DATE"] == earliest.get(r["INDEX_NAME"]),
        axis=1,
    )
    n_dropped = int(inception_excl_mask.sum())
    inception_dropped = df[inception_excl_mask].copy()
    inception_dropped["REASON"] = "inception_artifact"
    df = df[~inception_excl_mask].copy()
    log(f"     inception-day Exclusions discarded: {n_dropped}")

    # Dedupe
    df = df.drop_duplicates(subset=["INDEX_NAME", "EVENT_DATE", "SYMBOL", "EVENT_TYPE"])

    log(
        f"     resolved events: {len(df):,}  "
        f"(resolved_symbol={df['SYMBOL'].notna().sum():,}, unresolved={df['SYMBOL'].isna().sum():,})"
    )
    return df, inception_dropped


# =============================================================================
# 7. STAGE S4 — Reconstruct compositions per index
# =============================================================================
def _sym_set(series):
    """Return a set of uppercased symbols from a pandas Series."""
    return set(str(s).strip().upper() for s in series if s and not pd.isna(s))


def reconstruct(df_f1, df_f3, df_events):
    """
    Run per-index walker. Returns (df_compositions, df_events_unified, df_validation).
    """
    log("S4: reconstructing compositions per index")

    # Build snapshots: {index: {date: set(symbols)}}
    f1_snap = defaultdict(dict)
    for (idx, ts), g in df_f1.groupby(["INDEX_NAME", "TIME_STAMP"]):
        f1_snap[idx][ts] = _sym_set(g["SYMBOL"])
    f3_snap = defaultdict(dict)
    for (idx, ts), g in df_f3.groupby(["INDEX_NAME", "TIME_STAMP"]):
        f3_snap[idx][ts] = _sym_set(g["SYMBOL"])

    # Build events: {index: {date: (set_incls, set_excls)}}
    f2_ev = defaultdict(lambda: defaultdict(lambda: (set(), set())))
    for _, r in df_events.iterrows():
        if r["SYMBOL"] is None or pd.isna(r["SYMBOL"]):
            continue  # skip unresolved events
        idx, dt, sym, typ = r["INDEX_NAME"], r["EVENT_DATE"], r["SYMBOL"], r["EVENT_TYPE"]
        i_set, e_set = f2_ev[idx][dt]
        if typ == "INCL":
            i_set.add(sym)
        else:
            e_set.add(sym)
        f2_ev[idx][dt] = (i_set, e_set)

    # Enumerate indices to reconstruct: those with ANY snapshot (F1 or F3)
    anchored_indices = set(f1_snap.keys()) | set(f3_snap.keys())
    f2_indices = set(f2_ev.keys())
    orphans = f2_indices - anchored_indices
    if orphans:
        log(
            f"     skipping {len(orphans)} orphan F2 indices (no F1/F3 anchor): "
            f"{sorted(orphans)[:10]}{'...' if len(orphans) > 10 else ''}"
        )

    # Build F1 industry/cap-weight lookup
    # For each (INDEX, SYMBOL, DATE), we have INDUSTRY/CAP_WEIGHT in F1. Build lookup:
    f1_meta = {}  # (INDEX, SYMBOL) -> most recent {INDUSTRY, CAP_WEIGHT, date}
    for _, r in df_f1.iterrows():
        key = (r["INDEX_NAME"], r["SYMBOL"])
        d = r["TIME_STAMP"]
        if key not in f1_meta or d > f1_meta[key][0]:
            f1_meta[key] = (d, r.get("INDUSTRY"), r.get("CAP_WEIGHT"))

    def meta_for(index, symbol):
        rec = f1_meta.get((index, symbol))
        if rec is None:
            return (None, None)
        return (rec[1], rec[2])

    composition_rows = []
    event_rows = []
    validation_rows = []

    indices_sorted = sorted(anchored_indices)
    log(f"     reconstructing {len(indices_sorted)} indices")

    for index in indices_sorted:
        # Gather per-index data
        snapshots = {}
        source_of_snap = {}  # date -> 'F1_SNAPSHOT' or 'F3_SNAPSHOT'
        for d, s in f1_snap.get(index, {}).items():
            snapshots[d] = s
            source_of_snap[d] = "F1_SNAPSHOT"
        for d, s in f3_snap.get(index, {}).items():
            if d in snapshots and snapshots[d] != s:
                # Conflict between F1 and F3 on same date (shouldn't happen given date ranges)
                log(f"     [WARN] conflicting snapshots for {index} on {d}", "WARN")
            snapshots[d] = s
            source_of_snap[d] = source_of_snap.get(d, "F3_SNAPSHOT")

        events = f2_ev.get(index, {})

        if not snapshots:
            continue  # orphan (already logged)

        # Anchor = earliest snapshot
        anchor_date = min(snapshots.keys())
        anchor_state = snapshots[anchor_date]

        # ----------------- BACKWARD WALK -----------------
        pre_dates = sorted([d for d in events.keys() if d < anchor_date], reverse=True)
        state = set(anchor_state)
        for d in pre_dates:
            # state currently represents composition at d (no events between d and next date change)
            # Emit rows for this date
            for sym in sorted(state):
                industry, capw = meta_for(index, sym)
                composition_rows.append(
                    {
                        "INDEX_NAME": index,
                        "TIME_STAMP": d,
                        "SYMBOL": sym,
                        "INDUSTRY": industry,
                        "CAP_WEIGHT": capw,
                        "SOURCE": "RECONSTRUCTED",
                        "ANCHOR_DATE": anchor_date,
                        "CONFIDENCE": "high" if sym in anchor_state else "medium",
                    }
                )
            # Emit events for this date
            incls, excls = events[d]
            for sym in incls:
                event_rows.append(
                    {
                        "INDEX_NAME": index,
                        "EVENT_DATE": d,
                        "SYMBOL": sym,
                        "EVENT_TYPE": "INCL",
                        "SOURCE": "F2",
                    }
                )
            for sym in excls:
                event_rows.append(
                    {
                        "INDEX_NAME": index,
                        "EVENT_DATE": d,
                        "SYMBOL": sym,
                        "EVENT_TYPE": "EXCL",
                        "SOURCE": "F2",
                    }
                )
            # Reverse events to get state for the next earlier date
            for sym in incls:
                state.discard(sym)  # reverse inclusion = remove
            for sym in excls:
                state.add(sym)  # reverse exclusion = add back

        # ----------------- FORWARD WALK -----------------
        state = set(anchor_state)
        # Emit anchor date
        anchor_src = source_of_snap[anchor_date]
        for sym in sorted(state):
            industry, capw = meta_for(index, sym)
            composition_rows.append(
                {
                    "INDEX_NAME": index,
                    "TIME_STAMP": anchor_date,
                    "SYMBOL": sym,
                    "INDUSTRY": industry,
                    "CAP_WEIGHT": capw,
                    "SOURCE": anchor_src,
                    "ANCHOR_DATE": anchor_date,
                    "CONFIDENCE": "ground_truth",
                }
            )

        # All dates to visit post-anchor: union of F2 event dates and snapshot dates
        post_dates = sorted(
            set(d for d in events.keys() if d > anchor_date) | set(d for d in snapshots.keys() if d > anchor_date)
        )

        prev_snap_date = anchor_date  # used for GAP_DIFF detection

        for d in post_dates:
            # Apply F2 events at this date
            if d in events:
                incls, excls = events[d]
                for sym in excls:
                    state.discard(sym)
                for sym in incls:
                    state.add(sym)
                # Emit events
                for sym in incls:
                    event_rows.append(
                        {
                            "INDEX_NAME": index,
                            "EVENT_DATE": d,
                            "SYMBOL": sym,
                            "EVENT_TYPE": "INCL",
                            "SOURCE": "F2",
                        }
                    )
                for sym in excls:
                    event_rows.append(
                        {
                            "INDEX_NAME": index,
                            "EVENT_DATE": d,
                            "SYMBOL": sym,
                            "EVENT_TYPE": "EXCL",
                            "SOURCE": "F2",
                        }
                    )

            # Snap to snapshot if applicable
            if d in snapshots:
                truth = snapshots[d]
                missing = truth - state
                extra = state - truth
                if missing or extra:
                    # Classify: gap_diff if this is first snapshot after a long gap
                    kind = "snapshot_diff"
                    if (
                        source_of_snap[d] == "F3_SNAPSHOT"
                        and prev_snap_date
                        and source_of_snap.get(prev_snap_date) == "F1_SNAPSHOT"
                    ):
                        kind = "gap_diff"
                    validation_rows.append(
                        {
                            "INDEX_NAME": index,
                            "DATE": d,
                            "MISSING_FROM_RECON": ",".join(sorted(missing)),
                            "EXTRA_IN_RECON": ",".join(sorted(extra)),
                            "N_MISSING": len(missing),
                            "N_EXTRA": len(extra),
                            "TYPE": kind,
                        }
                    )
                    # Emit derived events to reconcile state to ground truth
                    derived_src = (
                        "GAP_DIFF"
                        if kind == "gap_diff"
                        else ("F1_DIFF" if source_of_snap[d] == "F1_SNAPSHOT" else "F3_DIFF")
                    )
                    for sym in missing:  # these should have been INCL but we didn't know
                        event_rows.append(
                            {
                                "INDEX_NAME": index,
                                "EVENT_DATE": d,
                                "SYMBOL": sym,
                                "EVENT_TYPE": "INCL",
                                "SOURCE": derived_src,
                            }
                        )
                    for sym in extra:  # these should have been EXCL but we didn't know
                        event_rows.append(
                            {
                                "INDEX_NAME": index,
                                "EVENT_DATE": d,
                                "SYMBOL": sym,
                                "EVENT_TYPE": "EXCL",
                                "SOURCE": derived_src,
                            }
                        )
                state = set(truth)
                src = source_of_snap[d]
                conf = "ground_truth"
                prev_snap_date = d
            else:
                src = "RECONSTRUCTED"
                conf = "high"

            # Emit composition rows
            for sym in sorted(state):
                industry, capw = meta_for(index, sym)
                composition_rows.append(
                    {
                        "INDEX_NAME": index,
                        "TIME_STAMP": d,
                        "SYMBOL": sym,
                        "INDUSTRY": industry,
                        "CAP_WEIGHT": capw,
                        "SOURCE": src,
                        "ANCHOR_DATE": anchor_date,
                        "CONFIDENCE": conf,
                    }
                )

    df_comp = pd.DataFrame(composition_rows)
    df_evt_unified = pd.DataFrame(event_rows).drop_duplicates(
        subset=["INDEX_NAME", "EVENT_DATE", "SYMBOL", "EVENT_TYPE", "SOURCE"]
    )
    df_val = pd.DataFrame(validation_rows)

    log(f"     compositions: {len(df_comp):,} rows")
    log(f"     events emitted: {len(df_evt_unified):,}")
    log(f"     validation discrepancies: {len(df_val)}")
    return df_comp, df_evt_unified, df_val


# =============================================================================
# 8. STAGE S5 — Validate (final summary; discrepancies already collected in S4)
# =============================================================================
def validate_and_report(df_comp, df_f1, df_f3, df_val):
    """Write validation_report.csv and print summary."""
    log("S5: writing validation report")
    if df_val.empty:
        pd.DataFrame(
            columns=["INDEX_NAME", "DATE", "MISSING_FROM_RECON", "EXTRA_IN_RECON", "N_MISSING", "N_EXTRA", "TYPE"]
        ).to_csv(OUT_VALIDATION, index=False)
    else:
        df_val.sort_values(["INDEX_NAME", "DATE"]).to_csv(OUT_VALIDATION, index=False, encoding="utf-8")

    # Cardinality sanity: check against expected sizes embedded in index name
    def expected_size(idx):
        m = re.findall(r"\b(\d{1,4})\b", idx)
        if not m:
            return None
        nums = [int(x) for x in m]
        # Prefer the last number (e.g., "NIFTY 500" -> 500; "NIFTY MIDCAP 150" -> 150)
        return nums[-1] if nums else None

    sizes = df_comp.groupby(["INDEX_NAME", "TIME_STAMP"])["SYMBOL"].nunique().reset_index()
    sizes["EXPECTED"] = sizes["INDEX_NAME"].map(expected_size)
    problematic = sizes[sizes["EXPECTED"].notna() & (sizes["SYMBOL"] != sizes["EXPECTED"])]
    log(f"     cardinality deviations: {len(problematic):,} (index,date) pairs differ from name-inferred size")

    log(f"     written: {OUT_VALIDATION.name} ({len(df_val)} rows)")


# =============================================================================
# 9. STAGE S6 — Refresh constituents.json (latest composition)
# =============================================================================
def refresh_constituents(df_comp):
    """Overwrite constituents.json with latest-date composition per index."""
    log("S6: refreshing constituents.json with latest composition")
    if df_comp.empty:
        log("     [WARN] empty compositions, skipping constituents.json refresh", "WARN")
        return

    latest_idx = df_comp.groupby("INDEX_NAME")["TIME_STAMP"].transform("max")
    latest = df_comp[df_comp["TIME_STAMP"] == latest_idx]

    out = {}
    for idx, g in latest.groupby("INDEX_NAME"):
        out[idx] = sorted(g["SYMBOL"].unique().tolist())

    # Preserve existing hand-curated entries that aren't in our reconstruction,
    # but overwrite reconstructed indices.
    existing = {}
    if OUT_CONSTITUENTS.exists():
        try:
            existing = json.loads(OUT_CONSTITUENTS.read_text("utf-8"))
        except Exception:
            pass

    # Merge: reconstructed wins; untouched indices preserved
    merged = dict(existing)
    merged.update(out)

    OUT_CONSTITUENTS.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    log(f"     wrote {len(out)} reconstructed indices (+{len(merged) - len(out)} preserved) " f"-> {OUT_CONSTITUENTS}")


# =============================================================================
# 10. MAIN
# =============================================================================
def main():
    print("=" * 72)
    print("  NSE Index Composition Reconstruction")
    print("=" * 72)

    # S1: Parse
    df_f1 = parse_f1()
    df_f2 = parse_f2()
    df_f3 = parse_f3()

    # S2: Build mapping
    symbol_universe = sorted(set(df_f1["SYMBOL"].unique()) | set(df_f3["SYMBOL"].unique()))
    scrip_names = sorted(df_f2["SCRIP_NAME"].unique())
    df_map = build_mapping(scrip_names, symbol_universe)

    # S3: Resolve events
    df_events, _df_inception = resolve_events(df_f2, df_map)

    # S4: Reconstruct
    df_comp, df_evt_unified, df_val = reconstruct(df_f1, df_f3, df_events)

    # Write outputs
    df_comp = df_comp.sort_values(["INDEX_NAME", "TIME_STAMP", "SYMBOL"]).reset_index(drop=True)
    df_comp.to_parquet(OUT_COMPOSITIONS, index=False, compression="snappy")
    log(f"     wrote {OUT_COMPOSITIONS.name} ({len(df_comp):,} rows)")

    df_evt_unified = df_evt_unified.sort_values(["INDEX_NAME", "EVENT_DATE", "SYMBOL"]).reset_index(drop=True)
    df_evt_unified.to_csv(OUT_EVENTS, index=False, encoding="utf-8")
    log(f"     wrote {OUT_EVENTS.name} ({len(df_evt_unified):,} rows)")

    # S5: Validate
    validate_and_report(df_comp, df_f1, df_f3, df_val)

    # S6: Refresh constituents.json
    refresh_constituents(df_comp)

    # Summary
    print("\n" + "=" * 72)
    print("  SUMMARY")
    print("=" * 72)
    print(f"  F1 snapshots:              {len(df_f1):,} rows, {df_f1['INDEX_NAME'].nunique()} indices")
    print(f"  F2 events (raw):           {len(df_f2):,}")
    print(f"  F3 snapshots:              {len(df_f3):,} rows, {df_f3['INDEX_NAME'].nunique()} indices")
    print(
        f"  Name-symbol map:           {len(df_map):,} names "
        f"(resolved={df_map['SYMBOL'].notna().sum():,}, unresolved={df_map['SYMBOL'].isna().sum():,})"
    )
    print(f"  Compositions (final):      {len(df_comp):,} rows")
    print(f"  Unique indices:            {df_comp['INDEX_NAME'].nunique()}")
    print(f"  Unique symbols:            {df_comp['SYMBOL'].nunique():,}")
    if not df_comp.empty:
        print(f"  Date range:                {df_comp['TIME_STAMP'].min()} to {df_comp['TIME_STAMP'].max()}")
    print(f"  Events emitted (unified):  {len(df_evt_unified):,}")
    print(f"  Validation discrepancies:  {len(df_val)}")
    if not df_val.empty:
        by_type = df_val["TYPE"].value_counts().to_dict()
        print(f"    by type:                 {by_type}")

    # Spot-checks
    print("\n  Spot-checks:")
    spot = df_comp[(df_comp["INDEX_NAME"] == "NIFTY 50") & (df_comp["TIME_STAMP"] == "2024-09-30")]
    print(f"    NIFTY 50 @ 2024-09-30:   {len(spot)} rows (expect 50)")
    spot2 = df_comp[(df_comp["INDEX_NAME"] == "NIFTY 50") & (df_comp["TIME_STAMP"] == "2020-03-31")]
    print(f"    NIFTY 50 @ 2020-03-31:   {len(spot2)} rows (expect 50)")

    print("\n  Outputs in:", DATA_DIR)
    print("    -", OUT_COMPOSITIONS.name)
    print("    -", OUT_EVENTS.name)
    print("    -", OUT_MAP.name)
    print("    -", OUT_UNRESOLVED.name)
    print("    -", OUT_VALIDATION.name)
    print("    -", OUT_OVERRIDES.name, "(edit to fix mapping errors)")
    print(f"    - {OUT_CONSTITUENTS}")
    print("=" * 72)


if __name__ == "__main__":
    main()
